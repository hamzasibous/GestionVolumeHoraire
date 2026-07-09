from django.shortcuts import render
from django.db import models

from .models import Comporte, Departement, Filiere, Module, Sceance, Vacation, Local, SemesterPeriod, ScheduleTask
from rest_framework import serializers, viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from users.models import Enseignant
import threading
import zipfile
import io
from .solver import run_genetic_algorithm
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from datetime import datetime, time, timedelta

def generate_timetable_pdf(filiere, semester):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    # --- Header ---
    p.setFont("Helvetica-Bold", 10)
    p.drawString(1*cm, height - 1*cm, "UNIVERSITE MOULAY ISMAIL")
    p.drawString(1*cm, height - 1.4*cm, "FACULTE DES SCIENCES")
    p.drawString(1*cm, height - 1.8*cm, "MEKNES")

    # --- Academic Year Calculation ---
    period = SemesterPeriod.objects.filter(semester=semester).first()
    academic_year = "2025-26"
    if period:
        start_year = period.date_debut.year
        start_month = period.date_debut.month
        if start_month >= 8: # Start of autumn semester
            academic_year = f"{start_year}-{(start_year + 1) % 100:02d}"
        else: # Spring semester
            academic_year = f"{start_year - 1}-{start_year % 100:02d}"

    p.drawRightString(width - 1*cm, height - 1*cm, f"Année universitaire :{academic_year}")
    p.drawRightString(width - 1*cm, height - 1.4*cm, f"Filière : {filiere.nom}")
    p.drawRightString(width - 1*cm, height - 1.8*cm, f"Semestre {semester}")
    p.drawRightString(width - 1*cm, height - 2.2*cm, "Section : A")

    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height - 3*cm, "EMPLOI DU TEMPS")
    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(width/2, height - 3.6*cm, "SEMESTRE DE PRINTEMPS" if semester in ['S2', 'S4', 'S6', 'M2', 'M4'] else "SEMESTRE D'AUTOMNE")

    # --- Grid Setup ---
    margin_left = 2*cm
    margin_top = 5*cm
    grid_width = width - 3*cm
    grid_height = height - 7*cm
    row_height = grid_height / 6
    col_width_day = 2.5*cm
    col_width_time = (grid_width - col_width_day) / 11

    days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
    times = ["8.30", "9.30", "10.30", "11.30", "12.30", "13.30", "14.30", "15.30", "16.30", "17.30", "18.30", "19.30"]

    p.setFont("Helvetica", 10)
    for i, t_label in enumerate(times):
        x = margin_left + col_width_day + i * col_width_time
        p.drawCentredString(x, height - margin_top + 0.5*cm, t_label)
        p.setDash(1, 2)
        p.line(x, height - margin_top, x, height - margin_top - grid_height)
        p.setDash()

    for i in range(7):
        y = height - margin_top - i * row_height
        p.line(margin_left, y, margin_left + grid_width, y)
        if i < 6:
            p.setFont("Helvetica-Bold", 10)
            p.drawString(margin_left + 0.2*cm, y - row_height/2, days[i])

    p.line(margin_left, height - margin_top, margin_left, height - margin_top - grid_height)
    p.line(margin_left + col_width_day, height - margin_top, margin_left + col_width_day, height - margin_top - grid_height)
    p.line(margin_left + grid_width, height - margin_top, margin_left + grid_width, height - margin_top - grid_height)

    # --- Fill Data ---
    sceances = Sceance.objects.filter(module__comporte__semestre=semester, module__comporte__filiere=filiere).select_related('module', 'local', 'enseignant')
    processed_slots = set()
    for s in sceances:
        if not s.heure_debut: continue
        day_idx = s.date.weekday()
        if day_idx > 5: continue
        slot_key = (day_idx, s.heure_debut.hour, s.heure_debut.minute, s.module_id, s.local_id)
        if slot_key in processed_slots: continue
        processed_slots.add(slot_key)
        start_minutes = s.heure_debut.hour * 60 + s.heure_debut.minute
        base_minutes = 8 * 60 + 30
        offset_minutes = start_minutes - base_minutes
        x_start = margin_left + col_width_day + (offset_minutes / 60) * col_width_time
        x_end = x_start + (s.duree / 60) * col_width_time
        y_top = height - margin_top - day_idx * row_height
        y_bottom = y_top - row_height
        y_mid = y_bottom + (y_top - y_bottom) / 2
        p.setStrokeColor(colors.black)
        p.setLineWidth(1)
        p.line(x_start + 5, y_mid, x_end - 5, y_mid)
        p.line(x_start + 5, y_mid, x_start + 10, y_mid + 3)
        p.line(x_start + 5, y_mid, x_start + 10, y_mid - 3)
        p.line(x_end - 5, y_mid, x_end - 10, y_mid + 3)
        p.line(x_end - 5, y_mid, x_end - 10, y_mid - 3)
        p.setFont("Helvetica-Bold", 7)
        mod_room_text = f"{s.module.nom} : {s.local}"
        if len(mod_room_text) > 40: mod_room_text = mod_room_text[:37] + "..."
        p.drawCentredString((x_start + x_end) / 2, y_mid + 6, mod_room_text)
        if s.enseignant:
            p.setFont("Helvetica-Oblique", 6)
            prof_text = f"Pr. {s.enseignant.nom} {s.enseignant.prenom}"
            p.drawCentredString((x_start + x_end) / 2, y_mid - 10, prof_text)

    p.setFont("Helvetica", 8)
    p.drawRightString(width - 1*cm, 1*cm, datetime.now().strftime("%d/%m/%Y"))
    p.showPage()
    p.save()
    pdf_data = buffer.getvalue()
    buffer.close()
    return pdf_data

class ExportTimetablePDFView(APIView):
    def get(self, request):
        filiere_id = request.query_params.get('filiere')
        semester = request.query_params.get('semester')
        if not filiere_id or not semester:
            return Response({"error": "Filiere and semester are required"}, status=400)
        try:
            filiere = Filiere.objects.get(id=filiere_id)
        except Filiere.DoesNotExist:
            return Response({"error": "Filiere not found"}, status=404)
        pdf_content = generate_timetable_pdf(filiere, semester)
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{filiere.nom}_{semester}.pdf"'
        return response

def generate_teacher_timetable_pdf(teacher, week_start_date):
    from .models import Sceance, SemesterPeriod
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    # --- Header ---
    p.setFont("Helvetica-Bold", 10)
    p.drawString(1*cm, height - 1*cm, "UNIVERSITE MOULAY ISMAIL")
    p.drawString(1*cm, height - 1.4*cm, "FACULTE DES SCIENCES")
    p.drawString(1*cm, height - 1.8*cm, "MEKNES")

    # --- Academic Year Calculation ---
    academic_year = "2025-26"
    period = SemesterPeriod.objects.filter(date_debut__lte=week_start_date, date_fin__gte=week_start_date).first()
    if not period:
        period = SemesterPeriod.objects.first()
    if period:
        start_year = period.date_debut.year
        start_month = period.date_debut.month
        if start_month >= 8:
            academic_year = f"{start_year}-{(start_year + 1) % 100:02d}"
        else:
            academic_year = f"{start_year - 1}-{start_year % 100:02d}"

    p.drawRightString(width - 1*cm, height - 1*cm, f"Année universitaire : {academic_year}")
    p.drawRightString(width - 1*cm, height - 1.4*cm, f"Enseignant : Pr. {teacher.nom} {teacher.prenom}")
    p.drawRightString(width - 1*cm, height - 1.8*cm, f"Semaine du : {week_start_date.strftime('%d/%m/%Y')}")

    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height - 3*cm, "EMPLOI DU TEMPS ENSEIGNANT")

    # --- Grid Setup ---
    margin_left = 2*cm
    margin_top = 5*cm
    grid_width = width - 3*cm
    grid_height = height - 7*cm
    row_height = grid_height / 6
    col_width_day = 2.5*cm
    col_width_time = (grid_width - col_width_day) / 11

    days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
    times = ["8.30", "9.30", "10.30", "11.30", "12.30", "13.30", "14.30", "15.30", "16.30", "17.30", "18.30", "19.30"]

    p.setFont("Helvetica", 10)
    for i, t_label in enumerate(times):
        x = margin_left + col_width_day + i * col_width_time
        p.drawCentredString(x, height - margin_top + 0.5*cm, t_label)
        p.setDash(1, 2)
        p.line(x, height - margin_top, x, height - margin_top - grid_height)
        p.setDash()

    for i in range(7):
        y = height - margin_top - i * row_height
        p.line(margin_left, y, margin_left + grid_width, y)
        if i < 6:
            p.setFont("Helvetica-Bold", 10)
            p.drawString(margin_left + 0.2*cm, y - row_height/2, days[i])

    p.line(margin_left, height - margin_top, margin_left, height - margin_top - grid_height)
    p.line(margin_left + col_width_day, height - margin_top, margin_left + col_width_day, height - margin_top - grid_height)
    p.line(margin_left + grid_width, height - margin_top, margin_left + grid_width, height - margin_top - grid_height)

    # --- Fill Data ---
    end_date = week_start_date + timedelta(days=6)
    sceances = Sceance.objects.filter(
        enseignant=teacher, 
        date__range=[week_start_date, end_date]
    ).select_related('module', 'local')
    
    processed_slots = set()
    for s in sceances:
        if not s.heure_debut: continue
        day_idx = s.date.weekday()
        if day_idx > 5: continue
        slot_key = (day_idx, s.heure_debut.hour, s.heure_debut.minute, s.module_id, s.local_id)
        if slot_key in processed_slots: continue
        processed_slots.add(slot_key)
        start_minutes = s.heure_debut.hour * 60 + s.heure_debut.minute
        base_minutes = 8 * 60 + 30
        offset_minutes = start_minutes - base_minutes
        x_start = margin_left + col_width_day + (offset_minutes / 60) * col_width_time
        x_end = x_start + (s.duree / 60) * col_width_time
        y_top = height - margin_top - day_idx * row_height
        y_bottom = y_top - row_height
        y_mid = y_bottom + (y_top - y_bottom) / 2
        p.setStrokeColor(colors.black)
        p.setLineWidth(1)
        p.line(x_start + 5, y_mid, x_end - 5, y_mid)
        p.line(x_start + 5, y_mid, x_start + 10, y_mid + 3)
        p.line(x_start + 5, y_mid, x_start + 10, y_mid - 3)
        p.line(x_end - 5, y_mid, x_end - 10, y_mid + 3)
        p.line(x_end - 5, y_mid, x_end - 10, y_mid - 3)
        p.setFont("Helvetica-Bold", 7)
        
        mod_room_text = f"{s.module.nom} : {s.local}"
        if len(mod_room_text) > 40: mod_room_text = mod_room_text[:37] + "..."
        p.drawCentredString((x_start + x_end) / 2, y_mid + 6, mod_room_text)
        
        # Show filiere
        filiere_names = list(s.module.filieres.values_list('nom', flat=True))
        filiere_text = f"Fil : {', '.join(filiere_names)}" if filiere_names else "Fil : N/A"
        p.setFont("Helvetica-Oblique", 6)
        p.drawCentredString((x_start + x_end) / 2, y_mid - 10, filiere_text)

    p.setFont("Helvetica", 8)
    p.drawRightString(width - 1*cm, 1*cm, datetime.now().strftime("%d/%m/%Y"))
    p.showPage()
    p.save()
    pdf_data = buffer.getvalue()
    buffer.close()
    return pdf_data

class ExportTeacherTimetablePDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        teacher = request.user
        week_start_str = request.query_params.get('week_start')
        if not week_start_str:
            return Response({"error": "week_start date is required"}, status=400)
        try:
            week_start_date = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format, use YYYY-MM-DD"}, status=400)

        pdf_content = generate_teacher_timetable_pdf(teacher, week_start_date)
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{teacher.nom}_{teacher.prenom}_{week_start_str}.pdf"'
        return response

class ExportAllTimetablesZIPView(APIView):
    def get(self, request):
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w') as zip_file:
            filieres = Filiere.objects.all()
            for filiere in filieres:
                semesters = Comporte.objects.filter(filiere=filiere).values_list('semestre', flat=True).distinct()
                for semester in semesters:
                    pdf_content = generate_timetable_pdf(filiere, semester)
                    clean_f_name = "".join([c for c in filiere.nom if c.isalnum() or c in (' ', '_', '-')]).strip()
                    zip_path = f"{clean_f_name}/{semester}.pdf"
                    zip_file.writestr(zip_path, pdf_content)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="tous_les_emplois_du_temps.zip"'
        return response


class FiliereSerializer(serializers.ModelSerializer):
    class Meta:
        model = Filiere
        fields = ["id", "nom", "description", "niveaux", "departement", "modules"]


class DepartementSerializer(serializers.ModelSerializer):
    class Meta:
        model = Departement
        fields = ["id", "nom"]


class ModuleSerializer(serializers.ModelSerializer):
    class Meta:
        model = Module
        fields = ["id", "nom"]


class ComporteSerlizer(serializers.ModelSerializer):
    class Meta:
        model = Comporte
        fields = ["id", "filiere", "module", "semestre"]


class LocalSerializer(serializers.ModelSerializer):
    name = serializers.CharField(source="__str__", read_only=True)
    departement_name = serializers.CharField(source="departement.nom", read_only=True)

    class Meta:
        model = Local
        fields = ["id", "bloc", "numero", "capacite", "is_amphi", "departement", "departement_name", "name"]


class LocalViewSet(viewsets.ModelViewSet):
    queryset = Local.objects.all()
    serializer_class = LocalSerializer


class SceanceSerializer(serializers.ModelSerializer):
    enseignant_name = serializers.SerializerMethodField()
    enseignant_email = serializers.CharField(source="enseignant.email", read_only=True, default="")
    enseignant_tel = serializers.CharField(source="enseignant.tel", read_only=True, default="")
    local_name = serializers.CharField(source="local.__str__", read_only=True)
    module_name = serializers.CharField(source="module.nom", read_only=True)
    
    class Meta:
        model = Sceance
        fields = ["id", "type", "duree", "date", "heure_debut", "module", "module_name", "enseignant", "enseignant_name", "enseignant_email", "enseignant_tel", "local", "local_name"]

    def get_enseignant_name(self, obj):
        if obj.enseignant:
            return str(obj.enseignant)
        return ""


class VacationSerializer(serializers.ModelSerializer):
    teacher_name = serializers.CharField(source="enseignant.__str__", read_only=True)

    class Meta:
        model = Vacation
        fields = ["id", "enseignant", "teacher_name", "titre", "date_debut", "date_fin", "type_conge", "statut", "is_global"]


class VacationViewSet(viewsets.ModelViewSet):
    queryset = Vacation.objects.all()
    serializer_class = VacationSerializer

    def create(self, request, *args, **kwargs):
        is_bulk = request.data.get('enseignant') == 'all' or request.data.get('is_global') == True
        
        if is_bulk:
            data = request.data.copy()
            data['is_global'] = True
            data['enseignant'] = None # Global holiday doesn't need a specific teacher
            if not data.get('titre'):
                data['titre'] = data.get('type_conge', 'Global Holiday')
            
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return super().create(request, *args, **kwargs)


class ModuleDetailSerializer(serializers.ModelSerializer):
    total_hours = serializers.SerializerMethodField()
    assigned_hours = serializers.SerializerMethodField()
    semestre = serializers.SerializerMethodField()
    v_h_hebdo = serializers.SerializerMethodField()
    seances = SceanceSerializer(many=True, source="sceances", read_only=True)
    
    class Meta:
        model = Module
        fields = ["id", "nom", "total_hours", "assigned_hours", "semestre", "v_h_hebdo", "seances"]

    def get_total_hours(self, obj):
        name_lower = obj.nom.lower()
        if 'pfe' in name_lower or 'projet de fin' in name_lower:
            return 0
        import re
        match = re.search(r'\(S\d+ - (\d+)h\)', obj.nom)
        if match:
            return int(match.group(1))
        # fallback
        vh = self.get_v_h_hebdo(obj)
        return 12 * vh

    def get_assigned_hours(self, obj):
        if hasattr(obj, 'prefetched_assigned_minutes'):
            return obj.prefetched_assigned_minutes / 60
        return sum(s.duree for s in obj.sceances.all()) / 60

    def _get_comporte(self, obj):
        filiere_id = self.context.get('filiere_id')
        if not filiere_id: return None
        if hasattr(obj, 'prefetched_comportes'):
            for c in obj.prefetched_comportes:
                if c.filiere_id == filiere_id: return c
        return Comporte.objects.filter(filiere_id=filiere_id, module=obj).first()

    def get_semestre(self, obj):
        c = self._get_comporte(obj)
        return c.semestre if c else None

    def get_v_h_hebdo(self, obj):
        c = self._get_comporte(obj)
        return c.v_h_hebdo if c else 0

class FiliereDetailSerializer(serializers.ModelSerializer):
    modules = serializers.SerializerMethodField()
    niveaux_display = serializers.CharField(source="get_niveaux_display", read_only=True)
    
    class Meta:
        model = Filiere
        fields = ["id", "nom", "description", "niveaux", "niveaux_display", "modules"]

    def get_modules(self, obj):
        # Use prefetched modules and comporta data
        modules = list(obj.modules.all())
        comportes = list(obj.comporte_set.all())
        for m in modules:
            m.prefetched_comportes = [c for c in comportes if c.module_id == m.id]
            m.prefetched_assigned_minutes = sum(s.duree for s in m.sceances.all())
        return ModuleDetailSerializer(modules, many=True, context={'filiere_id': obj.id}).data


# In your views.py
from rest_framework.generics import CreateAPIView, ListAPIView, RetrieveUpdateDestroyAPIView


from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from .ai_service import AIService

class ExtractVacationsView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file_obj = request.data.get('image')
        if not file_obj:
            return Response({"error": "No image provided"}, status=status.HTTP_400_BAD_REQUEST)

        ai_service = AIService()
        results = ai_service.extract_vacations_from_image(file_obj)
        
        if isinstance(results, dict) and "error" in results:
            return Response(results, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        return Response(results, status=status.HTTP_200_OK)


class ExtractTimetableView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file_obj = request.data.get('file')
        filiere_id = request.data.get('filiere_id')
        semester = request.data.get('semester')

        if not file_obj or not filiere_id or not semester:
            return Response({"error": "Missing file, filiere_id, or semester"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            filiere = Filiere.objects.get(id=filiere_id)
        except Filiere.DoesNotExist:
            return Response({"error": "Filiere not found"}, status=status.HTTP_404_NOT_FOUND)

        ai_service = AIService()
        results = ai_service.extract_timetable_from_image(file_obj)
        
        if isinstance(results, dict) and "error" in results:
            print(f"CRITICAL AI EXTRACTION ERROR: {results['error']}")
            return Response(results, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        from .models import Module, Local, Sceance, SemesterPeriod
        from users.models import Enseignant
        from datetime import datetime, time, timedelta
        import numpy as np

        period = SemesterPeriod.objects.filter(semester=semester).first()
        if not period:
            return Response({"error": f"Period for semester {semester} not found"}, status=400)

        available_modules = list(Module.objects.filter(comporte__filiere=filiere, comporte__semestre=semester))
        assigned_module_ids = Sceance.objects.filter(module__comporte__filiere=filiere, module__comporte__semestre=semester).values_list('module_id', flat=True).distinct()
        unassigned_modules = [m for m in available_modules if m.id not in assigned_module_ids]
        
        # CRITICAL FIX: Only include REAL active professors
        all_profs = list(Enseignant.objects.filter(
            role__icontains='ENSEIGNANT',
            is_active=True
        ).exclude(email='admin@gmail.com').exclude(nom__icontains='assign'))
        
        all_rooms = list(Local.objects.all())
        days_map = {'Lundi': 0, 'Mardi': 1, 'Mercredi': 2, 'Jeudi': 3, 'Vendredi': 4, 'Samedi': 5}

        # 1. Pass 1: Robust Module Matching (NO CREATION)
        # We strictly use available_modules for this semester
        initial_sessions = []
        
        # Helper to find best module match
        def find_best_module(extracted_name):
            if not extracted_name: return None
            extracted_name = extracted_name.lower()
            # 1. Exact or partial match
            match = next((m for m in available_modules if extracted_name in m.nom.lower() or m.nom.lower() in extracted_name), None)
            return match

        # Rule Enforcement: 7 modules total. 5 modules get 2 sessions, 2 modules get 1 session.
        # Total = 12 sessions.
        # Track usage: {module_id: {'CM': count, 'TD': count, 'TP': count, 'total': count}}
        module_usage = {m.id: {'CM': 0, 'TD': 0, 'TP': 0, 'total': 0} for m in available_modules}
        
        def get_fallback_module(session_type_hint):
            # We want 5 modules to hit 2 (1 CM + 1 TD/TP), and 2 modules to hit 1 (CM only).
            for m in available_modules:
                usage = module_usage[m.id]
                
                # If module has 0 sessions, it can take anything, but prefer CM first
                if usage['total'] == 0:
                    return m, 'CM'
                
                # If module has 1 session, it can take a 2nd session (TD/TP) IF we haven't hit the limit of 5 double-modules
                if usage['total'] == 1:
                    modules_with_2 = sum(1 for v in module_usage.values() if v['total'] >= 2)
                    if modules_with_2 < 5 and usage['TD'] == 0 and usage['TP'] == 0:
                        return m, session_type_hint if session_type_hint in ['TD', 'TP'] else 'TD'
                        
            # absolute fallback
            return available_modules[0] if available_modules else None, session_type_hint

        for session in results:
            mod_name = session.get('module_name')
            ai_session_type = session.get('type', 'CM')
            matched_module = find_best_module(mod_name)
            
            # If no match or if the matched module is "full" (>=2), use fallback
            if not matched_module or module_usage.get(matched_module.id, {}).get('total', 0) >= 2:
                matched_module, assigned_type = get_fallback_module(ai_session_type)
            else:
                # We have a matched module. Let's determine its type based on what it already has
                usage = module_usage[matched_module.id]
                if usage['total'] == 0:
                    assigned_type = 'CM'
                elif usage['total'] == 1:
                    modules_with_2 = sum(1 for v in module_usage.values() if v['total'] >= 2)
                    if modules_with_2 < 5:
                        # Give it the requested type if it doesn't have it, otherwise fallback to whatever is missing
                        if ai_session_type in ['TD', 'TP'] and usage[ai_session_type] == 0:
                            assigned_type = ai_session_type
                        else:
                            assigned_type = 'TD' if usage['CM'] > 0 else 'CM'
                    else:
                        # This module is forced to stay at 1. Fallback to another module.
                        matched_module, assigned_type = get_fallback_module(ai_session_type)
                else:
                    matched_module, assigned_type = get_fallback_module(ai_session_type)

            if matched_module:
                module_usage[matched_module.id]['total'] += 1
                # Ensure the key exists in case AI returns something weird
                if assigned_type not in module_usage[matched_module.id]:
                    module_usage[matched_module.id][assigned_type] = 0
                module_usage[matched_module.id][assigned_type] += 1
                session['type'] = assigned_type

            matched_teacher = None
            if session.get('teacher_name'):
                name = session['teacher_name'].lower()
                matched_teacher = next((t for t in all_profs if name in f"{t.prenom} {t.nom}".lower() or f"{t.prenom} {t.nom}".lower() in name), None)

            matched_room = None
            room_name = session.get('room_name')
            if room_name:
                name = room_name.lower()
                matched_room = next((r for r in all_rooms if name in str(r).lower() or str(r).lower() in name), None)
                if not matched_room:
                    # Keep auto-creation for Rooms as requested previously
                    parts = room_name.split('.')
                    bloc = parts[0] if len(parts) > 1 else "EXT"
                    numero = parts[1] if len(parts) > 1 else room_name
                    matched_room = Local.objects.create(bloc=bloc, numero=numero, capacite=40, is_amphi=('amphi' in room_name.lower()), departement=filiere.departement)
                    all_rooms.append(matched_room)

            initial_sessions.append({
                "day": session.get('day'),
                "time": session.get('time'),
                "duration": session.get('duration', 120),
                "type": session.get('type', 'CM'),
                "module_id": matched_module.id if matched_module else None,
                "module_name": matched_module.nom if matched_module else session.get('module_name'),
                "teacher_id": matched_teacher.id if matched_teacher else None,
                "teacher_name": str(matched_teacher) if matched_teacher else session.get('teacher_name'),
                "room_id": matched_room.id if matched_room else None,
                "room_name": str(matched_room) if matched_room else session.get('room_name'),
            })

        # 2. Pass 2: Hungarian-style Proactive Assignment with Overlap & Fairness
        final_sessions = []
        teacher_busy_slots = {} # slot_key -> set(teacher_ids)
        room_busy_slots = {}    # slot_key -> set(room_ids)
        teacher_module_assignments = {} # teacher_id -> set(module_ids)
        
        # TRACKER: teacher_id -> set(module_ids) for THIS filiere AND semester
        # Initialize from database to respect existing assignments
        teacher_module_assignments = {}
        db_assignments = Sceance.objects.filter(
            module__comporte__filiere=filiere,
            module__comporte__semestre=semester
        ).values_list('enseignant_id', 'module_id')
        
        for t_id, m_id in db_assignments:
            if t_id:
                if t_id not in teacher_module_assignments: teacher_module_assignments[t_id] = set()
                teacher_module_assignments[t_id].add(m_id)

        # Global workloads
        teacher_workloads = {t.id: Sceance.objects.filter(enseignant_id=t.id).count() * 2 for t in all_profs}

        for session in initial_sessions:
            day_name_raw = (session.get('day') or '').strip().capitalize()
            time_str_raw = (session.get('time') or '').strip().replace('h', ':')
            mod_id = session.get('module_id')
            duration = session.get('duration', 120)
            
            # Use normalized day name
            day_name = day_name_raw if day_name_raw in days_map else None
            
            if not day_name or not time_str_raw:
                final_sessions.append(session)
                continue
            
            slot_key = f"{day_name}-{time_str_raw}"
            if slot_key not in teacher_busy_slots: teacher_busy_slots[slot_key] = set()
            if slot_key not in room_busy_slots: room_busy_slots[slot_key] = set()

            django_week_day = days_map[day_name] + 2
            
            # Robust time parsing
            try:
                if ':' not in time_str_raw:
                    time_str_raw = f"{time_str_raw.zfill(2)}:00"
                start_time = datetime.strptime(time_str_raw, "%H:%M").time()
            except:
                final_sessions.append(session)
                continue
            
            # DURATION-AWARE OVERLAP CHECK
            sessions_on_day = Sceance.objects.filter(date__week_day=django_week_day).select_related('enseignant', 'local')
            start_dt = datetime.combine(datetime.today(), start_time)
            end_dt = start_dt + timedelta(minutes=duration)
            
            occupied_prof_ids = set()
            occupied_room_ids = set()
            for s in sessions_on_day:
                if not s.heure_debut: continue
                s_start_dt = datetime.combine(datetime.today(), s.heure_debut)
                s_end_dt = s_start_dt + timedelta(minutes=s.duree)
                if (start_dt < s_end_dt) and (s_start_dt < end_dt):
                    if s.enseignant_id: occupied_prof_ids.add(s.enseignant_id)
                    if s.local_id: occupied_room_ids.add(s.local_id)
            
            total_busy_profs = occupied_prof_ids | teacher_busy_slots[slot_key]
            total_busy_rooms = occupied_room_ids | room_busy_slots[slot_key]

            # 1. Proactive Teacher Assignment (Force even if string exists but ID doesn't)
            if not session['teacher_id']:
                available_profs = []
                for p in all_profs:
                    # Is free?
                    if p.id in total_busy_profs: continue
                    
                    # Not teaching another module in this filiere AND semester?
                    assigned_mods = teacher_module_assignments.get(p.id, set())
                    if not assigned_mods or mod_id in assigned_mods:
                        available_profs.append(p)

                if available_profs:
                    # Sort by current workload (fairness)
                    available_profs.sort(key=lambda p: teacher_workloads.get(p.id, 0))
                    best_p = available_profs[0]
                    session['teacher_id'] = best_p.id
                    session['teacher_name'] = str(best_p)
                else:
                    print(f"DEBUG: No available professors found for slot {slot_key}")
            
            # Update trackers
            if session['teacher_id']:
                teacher_busy_slots[slot_key].add(session['teacher_id'])
                if mod_id:
                    if session['teacher_id'] not in teacher_module_assignments: teacher_module_assignments[session['teacher_id']] = set()
                    teacher_module_assignments[session['teacher_id']].add(mod_id)
                teacher_workloads[session['teacher_id']] = teacher_workloads.get(session['teacher_id'], 0) + 2

            # 2. Proactive Room Assignment
            if not session['room_id']:
                available_rooms = [r for r in all_rooms if r.id not in total_busy_rooms]
                if available_rooms:
                    best_r = next((r for r in available_rooms if (session['type'] == 'CM' and r.is_amphi) or (session['type'] != 'CM' and not r.is_amphi)), available_rooms[0])
                    session['room_id'] = best_r.id
                    session['room_name'] = str(best_r)
            
            if session['room_id']: 
                room_busy_slots[slot_key].add(session['room_id'])

            # Metadata for frontend - ALWAYS INCLUDE the currently assigned ID in the lists
            # so the dropdown can display it as selected.
            available_profs_for_slot = [{"id": p.id, "name": str(p)} for p in all_profs if p.id not in total_busy_profs or p.id == session['teacher_id']]
            available_rooms_for_slot = [{"id": r.id, "name": str(r)} for r in all_rooms if r.id not in total_busy_rooms or r.id == session['room_id']]
            
            session['available_profs'] = available_profs_for_slot
            session['available_rooms'] = available_rooms_for_slot
            session['needs_review'] = not (session['teacher_id'] and session['room_id'])
            final_sessions.append(session)

        return Response({
            "sessions": final_sessions,
            "all_profs": [{"id": p.id, "name": str(p)} for p in all_profs],
            "all_rooms": [{"id": r.id, "name": str(r)} for r in all_rooms],
            "available_modules": [{"id": m.id, "name": m.nom} for m in available_modules]
        }, status=status.HTTP_200_OK)


from rest_framework.permissions import AllowAny

class FiliereCreateView(CreateAPIView):
    queryset = Filiere.objects.all()
    serializer_class = FiliereSerializer

class FiliereDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Filiere.objects.all()
    serializer_class = FiliereSerializer
    permission_classes = [AllowAny]
    authentication_classes = []


class DepartmentCreateView(CreateAPIView):
    queryset = Departement.objects.all()
    serializer_class = DepartementSerializer


class ModuleCreateView(CreateAPIView):
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer

class ModuleDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer
    permission_classes = [AllowAny]
    authentication_classes = []


class ComporteCreateView(CreateAPIView):
    queryset = Comporte.objects.all()
    serializer_class = ComporteSerlizer

class DashboardStatsView(APIView):
    def get(self, request):
        from .models import Filiere, Module, Sceance, Local
        from users.models import Enseignant
        from django.db.models import Sum, Count
        
        user = request.user
        filiere_id = None
        if user.is_authenticated and any(r in user.role for r in ['RESPONSABLE_FILIERE', 'UTILISATEUR']) and not any(r in user.role for r in ['ADMIN', 'CHEF_DEPARTEMENT']) and user.filiere_id:
            filiere_id = user.filiere_id
            
        if filiere_id:
            filieres_qs = Filiere.objects.filter(id=filiere_id)
            modules_qs = Module.objects.filter(filieres__id=filiere_id)
            profs_qs = Enseignant.objects.filter(
                role__icontains='ENSEIGNANT',
                sceances__module__filieres__id=filiere_id
            ).distinct().exclude(email='admin@gmail.com').exclude(nom__icontains='assign')
            rooms_qs = Local.objects.filter(sceances__module__filieres__id=filiere_id).distinct()
            sceances_qs = Sceance.objects.filter(module__filieres__id=filiere_id)
            from users.models import Utilisateur
            total_students = Utilisateur.objects.filter(role__icontains='UTILISATEUR', filiere_id=filiere_id).count()
        else:
            filieres_qs = Filiere.objects.all()
            modules_qs = Module.objects.all()
            profs_qs = Enseignant.objects.filter(
                role__icontains='ENSEIGNANT'
            ).exclude(email='admin@gmail.com').exclude(nom__icontains='assign')
            rooms_qs = Local.objects.all()
            sceances_qs = Sceance.objects.all()
            from users.models import Utilisateur
            total_students = Utilisateur.objects.filter(role__icontains='UTILISATEUR').count()
            
        total_filieres = filieres_qs.count()
        total_modules = modules_qs.count()
        total_profs = profs_qs.count()
        total_rooms = rooms_qs.count()
        
        # Total volume in EqTD hours
        total_equiv_hours = 0
        for s in sceances_qs.only('type', 'duree'):
            hours = s.duree / 60.0
            if s.type == 'CM':
                total_equiv_hours += hours * 1.5
            elif s.type == 'TP':
                total_equiv_hours += hours * 0.75
            else:
                total_equiv_hours += hours * 1.0
        total_hours = round(total_equiv_hours)
        
        # Breakdown by type
        breakdown = sceances_qs.values('type').annotate(count=Count('id'))
        total_sceances = sum(item['count'] for item in breakdown)
        
        type_stats = {
            'CM': 0, 'TD': 0, 'TP': 0
        }
        if total_sceances > 0:
            for item in breakdown:
                type_stats[item['type']] = round((item['count'] / total_sceances) * 100)

        # Recent filieres with EqTD
        filieres_data = []
        for f in filieres_qs[:5]:
            f_sceances = Sceance.objects.filter(module__filieres=f).only('type', 'duree')
            f_equiv_hours = 0
            for s in f_sceances:
                hours = s.duree / 60.0
                if s.type == 'CM':
                    f_equiv_hours += hours * 1.5
                elif s.type == 'TP':
                    f_equiv_hours += hours * 0.75
                else:
                    f_equiv_hours += hours * 1.0
            filieres_data.append({
                'name': f.nom,
                'level': f.get_niveaux_display(),
                'totalHours': round(f_equiv_hours),
                'status': 'Validated' # Placeholder
            })

        return Response({
            'total_filieres': total_filieres,
            'total_modules': total_modules,
            'total_profs': total_profs,
            'total_rooms': total_rooms,
            'total_hours': total_hours,
            'total_students': total_students,
            'type_stats': type_stats,
            'filieres': filieres_data,
            'avg_workload': (total_hours // total_profs) if total_profs > 0 else 0
        })

def get_session_hours(module_nom, session_type, all_session_types_for_module):
    import re
    if 'pfe' in module_nom.lower() or 'projet de fin' in module_nom.lower():
        return 0.0, 0.0

    match = re.search(r'\(S\d+ - (\d+)h\)', module_nom)
    total_raw = float(match.group(1)) if match else 48.0
    
    has_cm = 'CM' in all_session_types_for_module
    has_td_tp = any(t in all_session_types_for_module for t in ['TD', 'TP'])
    
    if has_cm and has_td_tp:
        raw_hours = total_raw / 2.0
    else:
        raw_hours = total_raw

    if session_type == 'CM':
        equiv_hours = raw_hours * 1.5
    elif session_type == 'TP':
        equiv_hours = raw_hours * 0.75
    else:
        equiv_hours = raw_hours * 1.0
        
    return raw_hours, equiv_hours


class MyAssignmentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from users.models import Utilisateur
        from .models import Sceance
        from django.db.models import Sum

        user = request.user
        
        # Group assignments by module and type
        assignments = []
        sceances_summary = Sceance.objects.filter(enseignant_id=user.id).values('module__id', 'module__nom', 'type').annotate(count=models.Count('id'))
        
        total_hours = 0
        total_cm_hours = 0
        total_td_hours = 0
        total_tp_hours = 0
        
        for item in sceances_summary:
            mod_id = item['module__id']
            mod_nom = item['module__nom']
            s_type = item['type']
            
            # Dynamically determine raw and EqTD hours
            module_session_types = list(Sceance.objects.filter(module_id=mod_id).values_list('type', flat=True).distinct())
            raw_hours, equiv_hours = get_session_hours(mod_nom, s_type, module_session_types)
            
            if s_type == 'CM':
                total_cm_hours += raw_hours
            elif s_type == 'TD':
                total_td_hours += raw_hours
            elif s_type == 'TP':
                total_tp_hours += raw_hours
            
            total_hours += equiv_hours
            
            first_sceance = Sceance.objects.filter(enseignant_id=user.id, module_id=mod_id, type=s_type).first()
            
            assignments.append({
                'id': f"m{mod_id}-{s_type}",
                'code': f"MOD{mod_id}",
                'name': mod_nom,
                'type': s_type,
                'hours': equiv_hours,
                'raw_hours': raw_hours,
                'students': 'Calculated', 
                'location': str(first_sceance.local) if first_sceance else 'N/A'
            })

        dept = user.departement
        seuil = dept.seuil_horaire if dept else 192

        return Response({
            'total_hours': total_hours,
            'statutory_requirement': seuil,
            'overload': max(0, total_hours - seuil),
            'breakdown': {
                'CM': total_cm_hours * 1.5,
                'TD': total_td_hours * 1.0,
                'TP': total_tp_hours * 0.75,
            },
            'assignments': assignments
        })

class FacultyAssignmentListView(APIView):
    def get(self, request):
        from users.models import Enseignant
        from .models import Module, Sceance
        from django.db.models import Sum, Count
        
        # Only fetch REAL active teachers
        faculty = Enseignant.objects.filter(is_active=True, role__icontains='ENSEIGNANT').exclude(email__icontains='admin').exclude(nom__icontains='assign').prefetch_related('modules', 'sceances')
        data = []
        
        for prof in faculty:
            assignments = []
            sceances_summary = prof.sceances.values('module__id', 'module__nom', 'type').annotate(count=Count('id'))
            
            total_hours = 0
            
            for item in sceances_summary:
                mod_id = item['module__id']
                mod_nom = item['module__nom']
                s_type = item['type']
                
                module_session_types = list(Sceance.objects.filter(module_id=mod_id).values_list('type', flat=True).distinct())
                raw_hours, equiv_hours = get_session_hours(mod_nom, s_type, module_session_types)
                
                total_hours += equiv_hours
                
                assignments.append({
                    'id': f"m{mod_id}-{s_type}",
                    'moduleCode': f"MOD{mod_id}",
                    'moduleName': mod_nom,
                    'type': s_type,
                    'hours': equiv_hours
                })

            dept = prof.departement
            seuil = dept.seuil_horaire if dept else 192

            data.append({
                'id': str(prof.id),
                'name': str(prof),
                'department': prof.departement.nom if prof.departement else 'N/A',
                'title': 'Professeur',
                'workload': total_hours,
                'maxWorkload': seuil,
                'isOverloaded': total_hours > seuil,
                'assignments': assignments,
                'initials': "".join([n[0] for n in str(prof).split() if n]).upper()[:2]
            })

        return Response(data)

class FiliereListView(ListAPIView):
    serializer_class = FiliereSerializer

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated and any(r in user.role for r in ['RESPONSABLE_FILIERE', 'UTILISATEUR']) and user.filiere_id:
            return Filiere.objects.filter(id=user.filiere_id)
        return Filiere.objects.all()

class DepartmentListView(ListAPIView):
    queryset = Departement.objects.all()
    serializer_class = DepartementSerializer

class LocalListView(ListAPIView):
    queryset = Local.objects.all()
    serializer_class = LocalSerializer

class SceanceViewSet(viewsets.ModelViewSet):
    queryset = Sceance.objects.all()
    serializer_class = SceanceSerializer

    def get_queryset(self):
        queryset = Sceance.objects.all().select_related('module', 'enseignant', 'local')
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        filiere_id = self.request.query_params.get('filiere')
        semester = self.request.query_params.get('semester')
        mine = self.request.query_params.get('mine')
        
        user = self.request.user

        if user.is_authenticated:
            # Strict scoping for student (UTILISATEUR)
            if 'UTILISATEUR' in user.role:
                if user.filiere_id:
                    queryset = queryset.filter(module__filieres__id=user.filiere_id)
                else:
                    return Sceance.objects.none()
            
            # Strict scoping for Responsable de Filiere
            elif 'RESPONSABLE_FILIERE' in user.role:
                if user.filiere_id:
                    if mine == 'true':
                        queryset = queryset.filter(module__filieres__id=user.filiere_id, enseignant=user)
                    else:
                        queryset = queryset.filter(module__filieres__id=user.filiere_id)
                elif filiere_id:
                    queryset = queryset.filter(module__filieres__id=filiere_id)
            
            # Others: teachers see their own if mine=true, admins see all
            else:
                if mine == 'true':
                    queryset = queryset.filter(enseignant=user)
                elif filiere_id:
                    queryset = queryset.filter(module__filieres__id=filiere_id)
        else:
            if filiere_id:
                queryset = queryset.filter(module__filieres__id=filiere_id)

        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        if semester:
            active_filiere = user.filiere_id if (user.is_authenticated and any(r in user.role for r in ['UTILISATEUR', 'RESPONSABLE_FILIERE']) and user.filiere_id) else filiere_id
            if active_filiere:
                queryset = queryset.filter(module__comporte__semestre=semester, module__comporte__filiere_id=active_filiere)

        return queryset

    def create(self, request, *args, **kwargs):
        from .models import Sceance, SemesterPeriod, Comporte
        from datetime import datetime, timedelta
        
        module_id = request.data.get('module')
        filiere_id = request.query_params.get('filiere') or request.data.get('filiere')
        semester = request.data.get('semester')
        initial_date_str = request.data.get('date')

        # LOGIC: If replace_existing is True, remove ALL sessions for this filiere/semester first
        if request.data.get('replace_existing'):
            if filiere_id and semester:
                Sceance.objects.filter(
                    module__comporte__filiere_id=filiere_id,
                    module__comporte__semestre=semester
                ).delete()
        
        # Validation: Ensure session date matches module semester period
        if module_id and initial_date_str:
            from .models import SemesterPeriod, Comporte
            from datetime import datetime
            
            # Find the semester for this module in the filiere
            comporte = Comporte.objects.filter(module_id=module_id, filiere_id=filiere_id, semestre=semester).first()
            if comporte:
                period = SemesterPeriod.objects.filter(semester=comporte.semestre).first()
                if period:
                    session_date = datetime.strptime(initial_date_str, '%Y-%m-%d').date()
                    if session_date < period.date_debut or session_date > period.date_fin:
                        return Response(
                            {"error": f"Cette date ({initial_date_str}) n'est pas comprise dans la période du {period.get_semester_display()} ({period.date_debut} à {period.date_fin})."},
                            status=status.HTTP_400_BAD_REQUEST
                        )

        number_of_sessions = int(request.data.get('number_of_sessions', 1))
        initial_date_str = request.data.get('date')
        if not initial_date_str:
            return Response({"error": "Date is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        from datetime import datetime, timedelta
        current_date = datetime.strptime(initial_date_str, '%Y-%m-%d').date()
        
        enseignant_id = request.data.get('enseignant')
        local_id = request.data.get('local')
        heure_debut_str = request.data.get('heure_debut')
        duree = int(request.data.get('duree', 0))
        
        if not heure_debut_str:
            return Response({"error": "Start time (heure_debut) is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        heure_debut = datetime.strptime(heure_debut_str, '%H:%M').time()
        
        created_sessions = []
        
        while len(created_sessions) < number_of_sessions:
            # Check for vacations and increment week if necessary
            on_vacation = Vacation.objects.filter(
                (models.Q(is_global=True) | models.Q(enseignant_id=enseignant_id if enseignant_id else None)),
                date_debut__lte=current_date,
                date_fin__gte=current_date
            ).exists()
            
            if on_vacation:
                current_date += timedelta(days=7)
                continue
            
            # Check for local conflicts
            start_dt = datetime.combine(current_date, heure_debut)
            end_dt = start_dt + timedelta(minutes=duree)
            
            # Find overlapping sessions in the same local
            existing_sessions = Sceance.objects.filter(date=current_date, local_id=local_id)
            conflict = False
            for sess in existing_sessions:
                if not sess.heure_debut: continue
                s_start = datetime.combine(sess.date, sess.heure_debut)
                s_end = s_start + timedelta(minutes=sess.duree)
                
                if (start_dt < s_end) and (s_start < end_dt):
                    conflict = True
                    break
            
            if conflict:
                current_date += timedelta(days=7)
                continue 
            
            # Create session
            data = request.data.copy()
            data['date'] = current_date.strftime('%Y-%m-%d')
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            session = serializer.save()
            created_sessions.append(serializer.data)
            
            # Increment for next session
            current_date += timedelta(days=7)

        return Response(created_sessions, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        date_str = request.data.get('date', instance.date.strftime('%Y-%m-%d'))
        heure_debut_str = request.data.get('heure_debut', instance.heure_debut.strftime('%H:%M') if instance.heure_debut else None)
        duree = int(request.data.get('duree', instance.duree))
        local_id = request.data.get('local', instance.local_id)
        enseignant_id = request.data.get('enseignant', instance.enseignant_id)

        from datetime import datetime, timedelta
        current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        heure_debut = datetime.strptime(heure_debut_str, '%H:%M').time() if heure_debut_str else None

        # 1. Vacation Check
        vacation_filter = models.Q(is_global=True)
        if enseignant_id:
            vacation_filter |= models.Q(enseignant_id=enseignant_id)
            
        on_vacation = Vacation.objects.filter(
            vacation_filter,
            date_debut__lte=current_date,
            date_fin__gte=current_date
        ).exists()
        
        if on_vacation:
            return Response({"error": "L'enseignant est en congé ou c'est un jour férié à cette date."}, status=status.HTTP_400_BAD_REQUEST)

        # 2. Conflict Check
        if heure_debut:
            start_dt = datetime.combine(current_date, heure_debut)
            end_dt = start_dt + timedelta(minutes=duree)
            
            # Exclude current instance from conflict check
            existing_sessions = Sceance.objects.filter(date=current_date, local_id=local_id).exclude(id=instance.id)
            for sess in existing_sessions:
                if not sess.heure_debut: continue
                s_start = datetime.combine(sess.date, sess.heure_debut)
                s_end = s_start + timedelta(minutes=sess.duree)
                
                if (start_dt < s_end) and (s_start < end_dt):
                    return Response({"error": f"Conflit de salle détecté avec une autre séance à {current_date} {heure_debut}"}, status=status.HTTP_400_BAD_REQUEST)

        return super().update(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def get_available_resources(self, request):
        date_str = request.query_params.get('date')
        heure_debut_str = request.query_params.get('heure_debut')
        duree_val = request.query_params.get('duree', 120)
        filiere_id = request.query_params.get('filiere_id')
        semester = request.query_params.get('semester')
        module_id = request.query_params.get('module_id')
        exclude_id = request.query_params.get('exclude_id')

        if not date_str or not heure_debut_str:
            return Response({"available_profs": [], "available_rooms": []})

        try:
            from datetime import datetime, timedelta
            from users.models import Enseignant
            from .models import Local, Sceance

            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            start_time = datetime.strptime(heure_debut_str, '%H:%M').time()
            duree = int(duree_val)
            start_dt = datetime.combine(date_obj, start_time)
            end_dt = start_dt + timedelta(minutes=duree)

            # 1. Find busy resources in this specific time window
            busy_sessions = Sceance.objects.filter(date=date_obj)
            if exclude_id and str(exclude_id).isdigit():
                busy_sessions = busy_sessions.exclude(id=exclude_id)
            
            occupied_prof_ids = set()
            occupied_room_ids = set()
            
            for s in busy_sessions:
                if not s.heure_debut: continue
                s_start = datetime.combine(s.date, s.heure_debut)
                s_end = s_start + timedelta(minutes=s.duree)
                if (start_dt < s_end) and (s_start < end_dt):
                    if s.enseignant_id: occupied_prof_ids.add(s.enseignant_id)
                    if s.local_id: occupied_room_ids.add(s.local_id)

            # 2. Get all real professors (filtered)
            all_profs = Enseignant.objects.filter(
                role__icontains='ENSEIGNANT',
                is_active=True
            ).exclude(email='admin@gmail.com').exclude(nom__icontains='assign')
            
            print(f"DEBUG: Found {all_profs.count()} total potential professors after role filters.")

            # 3. Filter available professors
            available_profs = []
            for p in all_profs:
                # HARD CONSTRAINT: Slot conflict (double booking)
                if p.id in occupied_prof_ids:
                    print(f"DEBUG: Prof {p.id} ({p.nom}) excluded due to slot conflict.")
                    continue
                
                # SOFT CONSTRAINT: One module per filiere per semester (for display/sorting)
                violates_filiere_rule = False
                if filiere_id and semester and module_id:
                    violates_filiere_rule = Sceance.objects.filter(
                        module__comporte__filiere_id=filiere_id,
                        module__comporte__semestre=semester,
                        enseignant_id=p.id
                    ).exclude(module_id=module_id).exists()
                
                available_profs.append({
                    "id": p.id,
                    "name": str(p),
                    "violates_rule": violates_filiere_rule
                })

            print(f"DEBUG: Returning {len(available_profs)} available professors.")

            # 4. Get available rooms
            available_rooms = Local.objects.exclude(id__in=occupied_room_ids)

            return Response({
                "available_profs": available_profs,
                "available_rooms": [{"id": r.id, "name": str(r)} for r in available_rooms]
            })
        except Exception as e:
            print(f"Error in get_available_resources: {str(e)}")
            return Response({"available_profs": [], "available_rooms": []})

    @action(detail=False, methods=['get'])
    def check_availability(self, request):
        local_id = request.query_params.get('local')
        date_str = request.query_params.get('date')
        heure_debut_str = request.query_params.get('heure_debut')
        duree = int(request.query_params.get('duree', 120))
        exclude_id = request.query_params.get('exclude_id')
        
        if not local_id or not date_str:
            return Response({"error": "Missing local or date"}, status=400)
            
        from datetime import datetime, timedelta
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        existing_sessions = Sceance.objects.filter(date=date, local_id=local_id)
        if exclude_id:
            existing_sessions = existing_sessions.exclude(id=exclude_id)
        
        if heure_debut_str:
            heure_debut = datetime.strptime(heure_debut_str, '%H:%M').time()
            start_dt = datetime.combine(date, heure_debut)
            end_dt = start_dt + timedelta(minutes=duree)
            
            conflicts = []
            for sess in existing_sessions:
                if not sess.heure_debut: continue
                s_start = datetime.combine(sess.date, sess.heure_debut)
                s_end = s_start + timedelta(minutes=sess.duree)
                
                if (start_dt < s_end) and (s_start < end_dt):
                    conflicts.append(sess)
            
            return Response({
                "available": len(conflicts) == 0,
                "conflicts": SceanceSerializer(conflicts, many=True).data
            })
            
        return Response({
            "available": not existing_sessions.exists(),
            "conflicts": SceanceSerializer(existing_sessions, many=True).data
        })


class GenerateScheduleView(APIView):
    def get(self, request):
        check_dates = request.query_params.get('check_dates')
        if check_dates == 'all':
            from .models import SemesterPeriod
            periods = SemesterPeriod.objects.all()
            return Response([{
                "semester": p.semester,
                "start": p.date_debut,
                "end": p.date_fin
            } for p in periods])
        return Response({"error": "Method not allowed"}, status=405)

    def post(self, request):
        action = request.data.get('action')
        if action == 'cancel':
            task_id = request.data.get('task_id')
            if not task_id:
                return Response({"error": "Task ID required"}, status=400)
            try:
                task = ScheduleTask.objects.get(id=task_id)
                if task.status == 'RUNNING':
                    task.status = 'CANCELLED'
                    task.message = "Annulation demandée..."
                    task.save()
                    return Response({"message": "Génération annulée"})
                return Response({"error": "Tâche non active"}, status=400)
            except ScheduleTask.DoesNotExist:
                return Response({"error": "Tâche non trouvée"}, status=404)

        semesters_str = request.data.get('semesters', '')
        if not semesters_str:
            return Response({"error": "No semesters provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        semester_codes = semesters_str.split(',')
        task = ScheduleTask.objects.create(status='RUNNING', progress=0, message="Démarrage de l'algorithme...")
        
        def bg_solver():
            try:
                def cancel_check():
                    task.refresh_from_db()
                    return task.status == 'CANCELLED'

                success, results = run_genetic_algorithm(semester_codes, task_id=task.id, cancel_check=cancel_check)
                
                # Re-check status
                task.refresh_from_db()
                if task.status == 'CANCELLED':
                    task.message = "Génération annulée par l'utilisateur."
                elif success:
                    task.status = 'COMPLETED'
                    task.result_data = results
                    task.message = "Génération terminée avec succès. Veuillez confirmer l'aperçu."
                else:
                    task.status = 'FAILED'
                    task.message = results
            except Exception as e:
                if str(e) == "CANCELLATION_REQUESTED":
                    task.status = 'CANCELLED'
                    task.message = "Génération annulée par l'utilisateur."
                else:
                    task.status = 'FAILED'
                    task.message = str(e)
            task.save()

        thread = threading.Thread(target=bg_solver)
        thread.start()
        return Response({"task_id": task.id, "message": "Génération démarrée"}, status=status.HTTP_202_ACCEPTED)


class ConfirmScheduleView(APIView):
    def post(self, request):
        task_id = request.data.get('task_id')
        if not task_id:
            return Response({"error": "Task ID required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            task = ScheduleTask.objects.get(id=task_id, status='COMPLETED')
            schedule = task.result_data
            semester_codes = list(set([s['semester'] for s in schedule]))
            periods = {p.semester: p for p in SemesterPeriod.objects.filter(semester__in=semester_codes)}

            from datetime import timedelta
            days_list = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
            times = ['08:30', '10:45', '14:30', '16:45']
            days_map = {'Lundi': 0, 'Mardi': 1, 'Mercredi': 2, 'Jeudi': 3, 'Vendredi': 4, 'Samedi': 5}

            semester_groups = {}
            for gene in schedule:
                sem = gene['semester']
                if sem not in semester_groups: semester_groups[sem] = []
                semester_groups[sem].append(gene)

            for sem, genes in semester_groups.items():
                if sem not in periods: continue
                period = periods[sem]
                
                # Delete existing sessions in this period for these semesters
                Sceance.objects.filter(module__comporte__semestre=sem, date__range=[period.date_debut, period.date_fin]).delete()

                for gene in genes:
                    slot_idx = gene['slot']
                    day_name = days_list[slot_idx // 4]
                    target_day_of_week = days_map[day_name]
                    
                    # Find the first valid date for this specific session
                    # Start from period.date_debut and find the first occurrence of target_day_of_week
                    curr_date = period.date_debut
                    while curr_date.weekday() != target_day_of_week:
                        curr_date += timedelta(days=1)
                    
                    sessions_created = 0
                    while sessions_created < 12:
                        # Check if the week containing curr_date is blocked
                        # Find the Monday of the week curr_date is in
                        monday_of_week = curr_date - timedelta(days=curr_date.weekday())
                        is_blocked = Vacation.objects.filter(
                            is_global=True, 
                            date_debut__lte=monday_of_week + timedelta(days=6), 
                            date_fin__gte=monday_of_week
                        ).count() >= 3
                        
                        if not is_blocked:
                            Sceance.objects.create(
                                type=gene.get('type', 'CM'), 
                                duree=120, 
                                date=curr_date, 
                                heure_debut=times[slot_idx % 4], 
                                module_id=gene['module_id'], 
                                enseignant_id=gene['teacher_id'], 
                                local_id=gene['room_id']
                            )
                            sessions_created += 1
                        
                        curr_date += timedelta(days=7)

            return Response({"message": "Emploi du temps enregistré avec succès !"}, status=status.HTTP_200_OK)
        except ScheduleTask.DoesNotExist:
            return Response({"error": "Task not found"}, status=status.HTTP_404_NOT_FOUND)


class TaskStatusView(APIView):
    permission_classes = [AllowAny]
    def get(self, request, task_id):
        try:
            task = ScheduleTask.objects.get(id=task_id)
            return Response({
                "id": task.id,
                "status": task.status,
                "progress": task.progress,
                "message": task.message
            }, status=status.HTTP_200_OK)
        except ScheduleTask.DoesNotExist:
            return Response({"error": "Task not found"}, status=status.HTTP_404_NOT_FOUND)


class ModuleListView(ListAPIView):
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer


class FiliereDetailListView(ListAPIView):
    serializer_class = FiliereDetailSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = Filiere.objects.all().prefetch_related(
            'modules__sceances__enseignant',
            'modules__sceances__local',
            'comporte_set'
        )
        if user.is_authenticated and any(r in user.role for r in ['RESPONSABLE_FILIERE', 'UTILISATEUR']) and user.filiere_id:
            return queryset.filter(id=user.filiere_id)
        return queryset


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from users.models import Role, Utilisateur, Enseignant

class SeuilHoraireView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        dept = user.departement
        seuil = dept.seuil_horaire if dept else 192
        return Response({"seuil_horaire": seuil})

    def put(self, request):
        user = request.user
        if not any(r in user.role for r in [Role.ADMIN, Role.CHEF_DEPARTEMENT]):
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)
        
        seuil = request.data.get("seuil_horaire")
        if seuil is None:
            return Response({"error": "seuil_horaire is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            seuil = int(seuil)
        except ValueError:
            return Response({"error": "seuil_horaire must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

        dept = user.departement
        if not dept:
            dept, _ = Departement.objects.get_or_create(nom="Informatique")
            user.departement = dept
            user.save()

        dept.seuil_horaire = seuil
        dept.save()
        return Response({"seuil_horaire": dept.seuil_horaire, "message": "Seuil horaire mis à jour avec succès"})


class ImportTeachersExcelView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        user = request.user
        if not any(r in user.role for r in [Role.ADMIN, Role.CHEF_DEPARTEMENT]):
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        file_obj = request.data.get('file')
        if not file_obj:
            return Response({"error": "Aucun fichier Excel fourni"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True)
            sheet = wb.active
            
            imported_count = 0
            updated_count = 0
            errors = []
            
            # Expect headers: Nom, Prénom, Email, Téléphone, Département, Rôle
            row_idx = 1
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_idx += 1
                if not row or not any(row):
                    continue
                
                nom = str(row[0]).strip() if row[0] is not None else ""
                prenom = str(row[1]).strip() if row[1] is not None else ""
                email = str(row[2]).strip() if row[2] is not None else ""
                tel = str(row[3]).strip() if row[3] is not None else ""
                dept_name = str(row[4]).strip() if row[4] is not None else "Informatique"
                role_str = str(row[5]).strip().upper() if len(row) > 5 and row[5] is not None else "ENSEIGNANT"
                spec_str = str(row[6]).strip().upper() if len(row) > 6 and row[6] is not None else "INFORMATIQUE"
                
                filiere_obj = None
                if len(row) > 7 and row[7] is not None:
                    filiere_name = str(row[7]).strip()
                    from core.models import Filiere
                    filiere_obj = Filiere.objects.filter(nom__iexact=filiere_name).first()

                if not email:
                    errors.append(f"Ligne {row_idx}: Email manquant")
                    continue
                if not nom or not prenom:
                    errors.append(f"Ligne {row_idx}: Nom ou Prénom manquant")
                    continue

                if role_str not in Role.values:
                    role_str = Role.ENSEIGNANT

                if spec_str not in ["INFORMATIQUE", "MATHEMATIQUES", "PHYSIQUE", "LANGUES", "AUTRE"]:
                    spec_str = "INFORMATIQUE"

                dept = None
                if dept_name:
                    dept, _ = Departement.objects.get_or_create(nom=dept_name)

                u, created = Utilisateur.objects.get_or_create(
                    email=email,
                    defaults={
                        'nom': nom,
                        'prenom': prenom,
                        'tel': tel,
                        'role': role_str,
                        'specialite': spec_str,
                        'departement': dept,
                        'filiere': filiere_obj,
                        'is_staff': (role_str in [Role.ADMIN, Role.CHEF_DEPARTEMENT, Role.RESPONSABLE_FILIERE])
                    }
                )
                
                if created:
                    u.set_password("Umi2024!")
                    u.save()
                    imported_count += 1
                else:
                    u.nom = nom
                    u.prenom = prenom
                    u.tel = tel
                    u.role = role_str
                    u.specialite = spec_str
                    u.departement = dept
                    if filiere_obj:
                        u.filiere = filiere_obj
                    u.is_staff = (role_str in [Role.ADMIN, Role.CHEF_DEPARTEMENT, Role.RESPONSABLE_FILIERE])
                    u.save()
                    updated_count += 1

            return Response({
                "message": f"{imported_count} enseignants importés, {updated_count} mis à jour.",
                "imported": imported_count,
                "updated": updated_count,
                "errors": errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"Erreur lors de la lecture du fichier Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class ExportWorkloadExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if not any(r in user.role for r in [Role.ADMIN, Role.CHEF_DEPARTEMENT, Role.RESPONSABLE_FILIERE]):
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        filiere_id = request.query_params.get('filiere_id')
        enseignant_id = request.query_params.get('enseignant_id')
        module_id = request.query_params.get('module_id')
        semester = request.query_params.get('semester')

        queryset = Sceance.objects.all().select_related('enseignant', 'module', 'local')

        if filiere_id:
            queryset = queryset.filter(module__comporte__filiere_id=filiere_id)
        if enseignant_id:
            queryset = queryset.filter(enseignant_id=enseignant_id)
        if module_id:
            queryset = queryset.filter(module_id=module_id)
        if semester:
            queryset = queryset.filter(module__comporte__semestre=semester)

        grouped_data = {}
        for s in queryset:
            prof = s.enseignant
            if not prof: continue
            
            prof_id = prof.id
            if prof_id not in grouped_data:
                grouped_data[prof_id] = {
                    'name': str(prof),
                    'department': prof.departement.nom if prof.departement else 'N/A',
                    'assignments': {}
                }
            
            key = (s.module_id, s.type)
            if key not in grouped_data[prof_id]['assignments']:
                module_session_types = list(Sceance.objects.filter(module_id=s.module_id).values_list('type', flat=True).distinct())
                raw_hours, equiv_hours = get_session_hours(s.module.nom, s.type, module_session_types)
                grouped_data[prof_id]['assignments'][key] = {
                    'module_name': s.module.nom,
                    'type': s.type,
                    'raw_hours': raw_hours,
                    'equiv_hours': equiv_hours
                }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Volume Horaire"

        headers = [
            "Enseignant", "Département", "Module", "Type d'enseignement", 
            "Heures Brutes", "Heures EqTD"
        ]
        
        title_font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        data_font = Font(name='Calibri', size=11)
        total_font = Font(name='Calibri', size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )
        
        ws.append(["RAPPORT DES VOLUMES HORAIRES - DEPARTEMENT D'INFORMATIQUE"])
        ws.merge_cells('A1:F1')
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        
        ws.append([])
        
        ws.append(headers)
        ws.row_dimensions[3].height = 24
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        row_num = 4
        total_raw = 0
        total_eqtd = 0
        
        for prof_id, prof_info in grouped_data.items():
            for key, asg in prof_info['assignments'].items():
                m_type = asg['type']
                raw_h = asg['raw_hours']
                eq_h = asg['equiv_hours']
                
                total_raw += raw_h
                total_eqtd += eq_h
                
                row_data = [
                    prof_info['name'],
                    prof_info['department'],
                    asg['module_name'],
                    m_type,
                    raw_h,
                    eq_h
                ]
                
                ws.append(row_data)
                
                for col_idx in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in [5, 6]:
                        cell.alignment = Alignment(horizontal='right')
                        cell.number_format = '0.00'
                    else:
                        cell.alignment = Alignment(horizontal='left')
                
                row_num += 1

        total_row = ["TOTAL", "", "", "", total_raw, total_eqtd]
        ws.append(total_row)
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = total_font
            cell.border = double_bottom_border
            if col_idx in [5, 6]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '0.00'
            else:
                cell.alignment = Alignment(horizontal='left')

        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_to_check = str(cell.value or '')
                if cell.row == 1: continue
                if len(val_to_check) > max_len:
                    max_len = len(val_to_check)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Rapport_Volume_Horaire.xlsx"
        return response
